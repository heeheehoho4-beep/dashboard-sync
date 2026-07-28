import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
import os
import time

def sync_excel_to_gsheets(excel_path: str, json_key_path: str, spreadsheet_url: str):
    print(f"[INFO] 엑셀 파일을 읽는 중: {excel_path}")
    
    # 1. 엑셀 워크북 읽기
    try:
        import shutil
        import os
        
        target_path = excel_path
        # Windows 환경 파일 잠금 회피를 위해 임시 파일로 복사 후 읽기
        if os.name == 'nt':
            import ctypes
            from ctypes import wintypes
            
            def win32_copy_locked_file(src_path, dst_path):
                GENERIC_READ = 0x80000000
                FILE_SHARE_READ = 0x00000001
                FILE_SHARE_WRITE = 0x00000002
                FILE_SHARE_DELETE = 0x00000004
                OPEN_EXISTING = 3
                FILE_ATTRIBUTE_NORMAL = 0x80
                INVALID_HANDLE_VALUE = -1
                
                kernel32 = ctypes.WinDLL('kernel32', use_last_error=True)
                
                CreateFileW = kernel32.CreateFileW
                CreateFileW.argtypes = [
                    wintypes.LPCWSTR, wintypes.DWORD, wintypes.DWORD, wintypes.LPVOID,
                    wintypes.DWORD, wintypes.DWORD, wintypes.HANDLE
                ]
                CreateFileW.restype = wintypes.HANDLE
                
                ReadFile = kernel32.ReadFile
                ReadFile.argtypes = [
                    wintypes.HANDLE, wintypes.LPVOID, wintypes.DWORD, wintypes.LPDWORD, wintypes.LPVOID
                ]
                ReadFile.restype = wintypes.BOOL
                
                CloseHandle = kernel32.CloseHandle
                CloseHandle.argtypes = [wintypes.HANDLE]
                CloseHandle.restype = wintypes.BOOL
                
                h_file = CreateFileW(
                    src_path,
                    GENERIC_READ,
                    FILE_SHARE_READ | FILE_SHARE_WRITE | FILE_SHARE_DELETE,
                    None,
                    OPEN_EXISTING,
                    FILE_ATTRIBUTE_NORMAL,
                    None
                )
                
                if h_file == INVALID_HANDLE_VALUE:
                    err = ctypes.get_last_error()
                    raise ctypes.WinError(err)
                    
                try:
                    chunk_size = 64 * 1024
                    buffer = ctypes.create_string_buffer(chunk_size)
                    bytes_read = wintypes.DWORD(0)
                    with open(dst_path, 'wb') as f_out:
                        while True:
                            success = ReadFile(h_file, buffer, chunk_size, ctypes.byref(bytes_read), None)
                            if not success:
                                err = ctypes.get_last_error()
                                raise ctypes.WinError(err)
                            if bytes_read.value == 0:
                                break
                            f_out.write(buffer.raw[:bytes_read.value])
                    return True
                finally:
                    CloseHandle(h_file)
            
            temp_path = "temp_sync_copy.xlsm"
            try:
                win32_copy_locked_file(excel_path, temp_path)
                target_path = temp_path
            except Exception as copy_err:
                print(f"[Auto-Sync] 임시 복사 실패: {copy_err}")
                pass # 실패하면 그냥 원본으로 시도
                
        # data_only=True로 수식의 결과값을 읽음
        wb = openpyxl.load_workbook(target_path, data_only=True, read_only=True)
        
        # 다 읽었으면 임시 파일 삭제 (선택 사항)
        if os.name == 'nt' and target_path == temp_path:
            try:
                pass # wb를 다 쓰고 나중에 지우거나, 덮어쓰기 되므로 그냥 둠
            except:
                pass
    except Exception as e:
        return False, f"엑셀 읽기 실패: {str(e)}"
        
    print("[INFO] 구글 시트에 연결하는 중...")
    # 2. 구글 API 인증
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    try:
        credentials = Credentials.from_service_account_file(json_key_path, scopes=scopes)
        client = gspread.authorize(credentials)
        sh = client.open_by_url(spreadsheet_url)
    except Exception as e:
        import traceback
        return False, f"구글 시트 인증 실패: {repr(e)}\n{traceback.format_exc()}\n(스프레드시트에 sheet-bot 이메일 편집자 권한이 있는지 확인하세요)"
        
    print("[INFO] 데이터를 동기화합니다...")
    # 3. 각 시트를 순회하며 업로드
    exclude_sheets = ["검색결과", "대시보드", "사용자계정", "Sheet1"]
    
    # 구글 시트에 현재 있는 워크시트 목록
    existing_worksheets = {ws.title: ws for ws in sh.worksheets()}
    
    success_count = 0
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name in exclude_sheets:
                continue
                
            ws_excel = wb[sheet_name]
            
            # 데이터 추출 (값이 있는 행만)
            data = []
            for row in ws_excel.iter_rows(values_only=True):
                clean_row = []
                is_empty = True
                for cell in row:
                    if cell is None:
                        clean_row.append("")
                    else:
                        is_empty = False
                        if hasattr(cell, 'strftime'):
                            clean_row.append(cell.strftime("%Y-%m-%d"))
                        else:
                            clean_row.append(str(cell))
                
                if not is_empty:
                    data.append(clean_row)
                    
            if not data:
                continue
                
            # 워크시트 준비
            if sheet_name in existing_worksheets:
                ws_gsheet = existing_worksheets[sheet_name]
            else:
                ws_gsheet = sh.add_worksheet(title=sheet_name, rows=str(max(len(data)+10, 100)), cols="26")
                existing_worksheets[sheet_name] = ws_gsheet
                
            # 데이터 덮어쓰기
            ws_gsheet.clear()
            ws_gsheet.update(values=data, range_name='A1')
            print(f"  -> '{sheet_name}' 시트 동기화 완료 ({len(data)}행)")
            success_count += 1
            
            # API 호출 제한 방지 (Too Many Requests 방지)
            time.sleep(1)
            
        # [NEW] 에셋(PDF, 이미지) 동기화 (system_assets 시트)
        print("[INFO] 에셋 파일(PDF/이미지)들을 동기화합니다...")
        import glob
        import base64
        
        assets_data = []
        # assets 폴더 내 schedule 로 시작하는 파일들 검색
        asset_files = []
        for ext in ('*.pdf', '*.png', '*.jpg', '*.jpeg'):
            asset_files.extend(glob.glob(os.path.join("assets", f"schedule{ext}")))
            
        for fpath in asset_files:
            fname = os.path.basename(fpath)
            try:
                with open(fpath, "rb") as f:
                    b64_str = base64.b64encode(f.read()).decode('utf-8')
                
                # 40000자씩 분할 (구글 시트 셀 용량 제한 5만자 우회)
                chunk_size = 40000
                chunks = [b64_str[i:i+chunk_size] for i in range(0, len(b64_str), chunk_size)]
                assets_data.append([fname] + chunks)
                print(f"  -> '{fname}' 인코딩 완료 (총 {len(chunks)}개 청크)")
            except Exception as e:
                print(f"  -> '{fname}' 읽기 실패: {e}")
                
        # system_assets 워크시트 가져오기 또는 생성
        if "system_assets" in existing_worksheets:
            ws_assets = existing_worksheets["system_assets"]
        else:
            ws_assets = sh.add_worksheet(title="system_assets", rows="20", cols="26")
        
        # 기존 데이터 무조건 초기화 (로컬에 파일이 없으면 시트도 비워야 함)
        ws_assets.clear()
        
        if assets_data:
            # 최대 컬럼 수 계산
            max_cols = max(len(row) for row in assets_data)
            # update 할 때 워크시트 컬럼이 부족하면 에러가 날 수 있으므로 resize
            ws_assets.resize(rows=max(len(assets_data)+5, 20), cols=max(max_cols + 5, 26))
            ws_assets.update(values=assets_data, range_name='A1')
            
        print(f"  -> 'system_assets' 에셋 데이터 업로드 완료")
            
        return True, f"✅ 총 {success_count}개의 데이터 시트 및 에셋 파일이 구글 스프레드시트에 성공적으로 동기화되었습니다!"
        
    except Exception as e:
        return False, f"동기화 중 오류 발생: {str(e)}"

if __name__ == "__main__":
    # 이 스크립트 단독 실행용 테스트
    excel_path = "★전략3지원센터_생보 위촉일정확인_v1.xlsm"
    if not os.path.exists(excel_path):
        excel_path = "temp_test.xlsm"
        if not os.path.exists(excel_path):
            excel_path = "appointment_data.xlsx"
            
    json_path = "dashboard-sync-503805-3ac60a044531.json"
    url = "https://docs.google.com/spreadsheets/d/1jWwDgw0rEGeb0R4_d1UsfOnXsfwbOjtu9BBiPTZ16p0/edit?gid=0#gid=0"
    
    success, msg = sync_excel_to_gsheets(excel_path, json_path, url)
    print(msg)
