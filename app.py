import streamlit as st
import pandas as pd
import os
import time
import base64
from datetime import datetime, timezone, timedelta

# ─────────────────────────────────────────────────────────────
#  설정 (본인 환경에 맞게 수정하세요)
# ─────────────────────────────────────────────────────────────
# 원드라이브 동기화 폴더에서 사용자의 파일을 순차적으로 검색합니다.
EXCEL_PATHS_TO_TRY = [
    os.environ.get("EXCEL_FILE_PATH", ""),
    r"C:\Users\User\OneDrive - 인카금융서비스\전략3지원센터\★전략3지원센터_생보 위촉일정확인_v1.xlsm",
    r"C:\Users\User\OneDrive - 인카금융서비스\함경희\★전략3지원센터_생보 위촉일정확인_v1.xlsm",
    r"C:\Users\User\OneDrive - 인카금융서비스\함경희\전략3지원센터_생보 위촉일정확인_v1.xlsm",
    "appointment_data.xlsx"
]

EXCEL_PATH = "appointment_data.xlsx"
for path in EXCEL_PATHS_TO_TRY:
    if path and os.path.exists(path):
        EXCEL_PATH = path
        break

# ─────────────────────────────────────────────────────────────
#  백그라운드 자동 동기화 (Auto-Sync) 스레드
# ─────────────────────────────────────────────────────────────
import threading

@st.cache_resource
def start_auto_sync_watcher(excel_path):
    class SyncWatcher:
        def __init__(self, path):
            self.path = path
            self.last_mtime = 0
            if os.path.exists(path):
                self.last_mtime = os.path.getmtime(path)
            self.thread = threading.Thread(target=self._watch, daemon=True)
            self.thread.start()
            
        def _watch(self):
            import sync_to_gsheets
            json_file = "dashboard-sync-503805-3ac60a044531.json"
            gsheet_url = "https://docs.google.com/spreadsheets/d/1jWwDgw0rEGeb0R4_d1UsfOnXsfwbOjtu9BBiPTZ16p0/edit?gid=0#gid=0"
            
            while True:
                time.sleep(60) # 1분 단위로 검사
                if os.path.exists(self.path):
                    current_mtime = os.path.getmtime(self.path)
                    if current_mtime > self.last_mtime:
                        print(f"[Auto-Sync] 엑셀 파일 변경 감지! 10초 대기 중...")
                        time.sleep(10) # 엑셀 저장이 완전히 끝날 때까지 10초 대기 (파일 잠금 방지)
                        
                        try:
                            print(f"[Auto-Sync] 자동 동기화 시작...")
                            sync_to_gsheets.sync_excel_to_gsheets(self.path, json_file, gsheet_url)
                            print(f"[Auto-Sync] 자동 동기화 성공!")
                        except Exception as e:
                            print(f"[Auto-Sync] 동기화 중 오류 발생: {e}")
                            
                        # 다음 감시를 위해 마지막 시간 갱신 (저장 도중 바뀌었을 수 있으므로 다시 읽음)
                        if os.path.exists(self.path):
                            self.last_mtime = os.path.getmtime(self.path)

    return SyncWatcher(excel_path)

# 대시보드가 켜질 때 한 번만 전역 스레드 실행 (로컬 환경일 경우에만)
IS_LOCAL = os.path.exists(EXCEL_PATH)
_watcher = start_auto_sync_watcher(EXCEL_PATH) if IS_LOCAL else None

# 결과 시트에서 제외할 시트 이름 목록 (검색 대상에서 제외)
EXCLUDE_SHEETS = ["검색결과", "대시보드", "사용자계정", "Sheet1"]

# VBA 코드 기준 추출 컬럼 인덱스 (0-based)
#   A=0(보험사), F=5(▶), G=6(차수), H=7(URL동의기한), I=8(위촉예정일), J=9(미위촉시 다음접수일정)
COL_INDICES = [0, 5, 6, 7, 8, 9]
COL_HEADERS = ["보험사", "▶", "차수", "URL 동의기한", "위촉예정일", "미위촉시 다음접수일정"]

# 캐시 갱신 주기 (초)
CACHE_TTL = 300   # 5분

# ─────────────────────────────────────────────────────────────
#  로그인 계정 설정 (사용자 ID / 비밀번호 / 부서)
#  실제 운영 시에는 별도 DB나 환경변수로 관리하세요.
# ─────────────────────────────────────────────────────────────
USERS = {
    "3center": {"password": "incar33**", "name": "전략3지원센터", "dept": "ALL", "role": "user"},
    "3jiwon": {"password": "incar33**", "name": "관리자(3지원)", "dept": "ALL", "role": "admin"}
}

# ─────────────────────────────────────────────────────────────
#  엑셀 데이터 로드 (캐싱)
# ─────────────────────────────────────────────────────────────
import openpyxl
import ctypes
from ctypes import wintypes

def win32_copy_locked_file(src_path, dst_path):
    """Windows API CreateFileW를 직접 호출하여 Excel 등이 독점 잠금(Lock)한 파일도 우회 복사합니다."""
    GENERIC_READ = 0x80000000
    FILE_SHARE_READ = 0x00000001
    FILE_SHARE_WRITE = 0x00000002
    FILE_SHARE_DELETE = 0x00000004
    OPEN_EXISTING = 3
    FILE_ATTRIBUTE_NORMAL = 0x80
    INVALID_HANDLE_VALUE = -1

    kernel32 = ctypes.WinDLL('kernel32', use_last_error=True)

    CreateFileW = kernel32.CreateFileW
    CreateFileW.argtypes = [
        wintypes.LPCWSTR, wintypes.DWORD, wintypes.DWORD, wintypes.LPVOID,
        wintypes.DWORD, wintypes.DWORD, wintypes.HANDLE
    ]
    CreateFileW.restype = wintypes.HANDLE

    ReadFile = kernel32.ReadFile
    ReadFile.argtypes = [
        wintypes.HANDLE, wintypes.LPVOID, wintypes.DWORD, wintypes.LPDWORD, wintypes.LPVOID
    ]
    ReadFile.restype = wintypes.BOOL

    CloseHandle = kernel32.CloseHandle
    CloseHandle.argtypes = [wintypes.HANDLE]
    CloseHandle.restype = wintypes.BOOL

    h_file = CreateFileW(
        src_path,
        GENERIC_READ,
        FILE_SHARE_READ | FILE_SHARE_WRITE | FILE_SHARE_DELETE,
        None,
        OPEN_EXISTING,
        FILE_ATTRIBUTE_NORMAL,
        None
    )
    
    if h_file == INVALID_HANDLE_VALUE:
        err = ctypes.get_last_error()
        raise ctypes.WinError(err)
        
    try:
        chunk_size = 64 * 1024  # 64KB
        buffer = ctypes.create_string_buffer(chunk_size)
        bytes_read = wintypes.DWORD(0)
        
        with open(dst_path, 'wb') as f_out:
            while True:
                success = ReadFile(h_file, buffer, chunk_size, ctypes.byref(bytes_read), None)
                if not success:
                    err = ctypes.get_last_error()
                    raise ctypes.WinError(err)
                if bytes_read.value == 0:
                    break
                f_out.write(buffer.raw[:bytes_read.value])
        return True
    finally:
        CloseHandle(h_file)

@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def load_excel_data(path: str):
    """엑셀 파일의 모든 시트를 메모리 절약형(read_only) 방식으로 안전하고 빠르게 로드합니다."""
    if not os.path.exists(path):
        return None, f"엑셀 파일을 찾을 수 없습니다: {path}"
    
    # 엑셀 파일이 열려있어 락(Lock)이 걸린 경우를 대비해 임시 사본을 생성하여 읽습니다.
    temp_path = "temp_locked_excel.xlsm"
    try:
        # Windows API를 호출해 락을 강제로 우회하여 복사
        win32_copy_locked_file(path, temp_path)
        
        # 복사본이 정상적으로 존재하면 복사본을 읽고, 실패했다면 원본 직접 읽기 시도
        read_path = temp_path if os.path.exists(temp_path) else path
        
        wb = openpyxl.load_workbook(read_path, read_only=True)
        sheets = {}
        
        for name in wb.sheetnames:
            if name not in EXCLUDE_SHEETS:
                ws = wb[name]
                data = []
                consecutive_empty = 0
                
                # 행 데이터를 하나씩 읽으며 빈 영역을 스킵하고 데이터 끝을 감지
                for row in ws.iter_rows(values_only=True):
                    # 행 전체가 비어있거나 공백인 경우
                    if all(v is None or str(v).strip() == "" for v in row):
                        consecutive_empty += 1
                        if consecutive_empty >= 20:  # 20줄 연속으로 데이터가 없으면 시트 종료로 간주
                            break
                        continue
                    
                    consecutive_empty = 0
                    data.append(row)
                
                if data:
                    sheets[name] = pd.DataFrame(data)
                else:
                    sheets[name] = pd.DataFrame()
                    
        wb.close()
        
        # 사용이 완료된 임시 복사본 파일 제거 시도
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except:
            pass
            
        return sheets, None
    except Exception as e:
        return None, f"파일 읽기 오류: {e}"

# ─────────────────────────────────────────────────────────────
#  클라우드 환경용 데이터 로드 (캐싱)
# ─────────────────────────────────────────────────────────────
@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def load_gsheets_data():
    """클라우드 환경에서 구글 스프레드시트 데이터를 로드합니다."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        
        # Streamlit secrets에서 가져오기 (dict 형태)
        if "gcp_service_account" not in st.secrets:
            return None, "Streamlit Secrets에 gcp_service_account가 설정되지 않았습니다."
            
        creds_dict = dict(st.secrets["gcp_service_account"])
        
        # TOML 파싱 중 \n 이 이스케이프 문자 그대로 들어간 경우를 위한 안전 장치
        if "private_key" in creds_dict:
            creds_dict["private_key"] = creds_dict["private_key"].replace('\\n', '\n')
            
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(credentials)
        
        spreadsheet_url = "https://docs.google.com/spreadsheets/d/1jWwDgw0rEGeb0R4_d1UsfOnXsfwbOjtu9BBiPTZ16p0/edit?gid=0#gid=0"
        sh = client.open_by_url(spreadsheet_url)
        
        sheets = {}
        for ws in sh.worksheets():
            name = ws.title
            
            # [NEW] 에셋 (PDF/이미지) 동기화 처리
            if name == "system_assets":
                assets_data = ws.get_all_values()
                if assets_data:
                    import base64
                    import glob
                    os.makedirs("assets", exist_ok=True)
                    
                    # 기존 캐시된 schedule 파일들 먼저 삭제 (오래된 깃허브 파일이나 지난달 파일)
                    for old_file in glob.glob(os.path.join("assets", "schedule*")):
                        try: os.remove(old_file)
                        except: pass
                        
                    for row in assets_data:
                        if not row or not row[0]: continue
                        fname = row[0]
                        # 빈 셀 제거 후 이어붙이기
                        b64_str = "".join([chunk for chunk in row[1:] if chunk])
                        try:
                            with open(os.path.join("assets", fname), "wb") as f:
                                f.write(base64.b64decode(b64_str))
                        except Exception as e:
                            print(f"Error decoding asset {fname}: {e}")
                continue
                
            if name not in EXCLUDE_SHEETS:
                data = ws.get_all_values()
                if data:
                    sheets[name] = pd.DataFrame(data)
                else:
                    sheets[name] = pd.DataFrame()
                    
        return sheets, None
    except Exception as e:
        import traceback
        return None, f"구글 시트 읽기 오류: {e}\n{traceback.format_exc()}"



# ─────────────────────────────────────────────────────────────
#  VBA 로직 변환: 다중 시트 검색 (마지막 행 우선, 역순 탐색)
# ─────────────────────────────────────────────────────────────
def search_all_sheets(sheets: dict, keyword: str, user_dept: str, target_companies=None):
    """
    VBA의 FindAndExtractData_CustomHeader_LastOnly 로직을 파이썬으로 구현.
    - 각 시트에서 아래→위 방향으로 탐색하여 가장 마지막(최신) 행만 추출.
    - 사용자 부서 필터링 적용 (ALL 이면 전체 조회 허용).
    - target_companies 리스트가 주어지면 해당 보험사 시트만 탐색.
    """
    results = []
    keyword_lower = keyword.strip().lower()

    for sheet_name, df in sheets.items():
        if sheet_name in ["notice", "requests", "system_assets"]:
            continue
            
        if target_companies and sheet_name not in target_companies:
            continue
            
        found_row = None

        match_count = 0
        
        # 역순 탐색 (VBA: For rowNum = lastRow To 1 Step -1)
        for row_idx in range(len(df) - 1, -1, -1):
            row = df.iloc[row_idx]
            row_str = " ".join([str(v) for v in row.values if pd.notna(v)]).lower()

            if keyword_lower in row_str:
                if found_row is None:
                    found_row = row
                match_count += 1

        if found_row is not None:
            # 컬럼 인덱스 범위 초과 방지
            extracted = []
            for idx in COL_INDICES:
                val = found_row.iloc[idx] if idx < len(found_row) else ""
                # 날짜 포맷 처리
                if isinstance(val, (pd.Timestamp, datetime)):
                    val = val.strftime("%Y-%m-%d")
                extracted.append("" if pd.isna(val) else str(val))

            # 부서 필터링: 추출 데이터 어딘가에 부서 정보가 있으면 비교
            # (부서 정보가 없을 경우 ALL 사용자에게만 노출)
            dept_matched = True
            if user_dept != "ALL":
                row_full_str = " ".join([str(v) for v in found_row.values if pd.notna(v)])
                if user_dept not in row_full_str:
                    dept_matched = False

            if dept_matched:
                row_data = {}
                for h, v in zip(COL_HEADERS, extracted):
                    row_data[h] = v
                
                # A열(보험사)이 비어있는 경우를 대비해 시트 이름으로 강제 덮어쓰기
                row_data["보험사"] = sheet_name
                
                row_data["재접수횟수"] = f"{match_count - 1}회" if match_count > 1 else "0회"
                results.append(row_data)

    return results


# ─────────────────────────────────────────────────────────────
#  세션 초기화
# ─────────────────────────────────────────────────────────────
def init_session():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
        st.session_state.user_id = ""
        st.session_state.user_name = ""
        st.session_state.user_dept = ""
        st.session_state.user_role = ""


# ─────────────────────────────────────────────────────────────
#  로그인 화면
# ─────────────────────────────────────────────────────────────
def render_login():
    st.markdown("""
    <div style="display:flex; justify-content:center; margin-top:80px;">
        <div style="width:100%; max-width:440px;">
    """, unsafe_allow_html=True)

    st.markdown("<h2 style='text-align:center; color:#4A90D9; margin-bottom:8px;'>🔐 생보 위촉일정 조회 시스템</h2>", unsafe_allow_html=True)
    st.markdown("<p style='text-align:center; color:#888; margin-bottom:32px;'>권한이 있는 사용자만 접근할 수 있습니다.</p>", unsafe_allow_html=True)

    with st.form("login_form"):
        user_id = st.text_input("아이디", placeholder="아이디를 입력하세요")
        password = st.text_input("비밀번호", type="password", placeholder="비밀번호를 입력하세요")
        submitted = st.form_submit_button("로그인", use_container_width=True)

        if submitted:
            if user_id in USERS and USERS[user_id]["password"] == password:
                st.session_state.logged_in = True
                st.session_state.user_id = user_id
                st.session_state.user_name = USERS[user_id]["name"]
                st.session_state.user_dept = USERS[user_id]["dept"]
                st.session_state.user_role = USERS[user_id].get("role", "user")
                st.rerun()
            else:
                st.error("아이디 또는 비밀번호가 올바르지 않습니다.")

    st.markdown("</div></div>", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
#  메인 대시보드 화면
# ─────────────────────────────────────────────────────────────
def render_dashboard():
    # 상단 헤더
    col1, col2, col3 = st.columns([6, 2, 2])
    with col1:
        st.markdown(f"<h2 style='color:#4A90D9; margin-bottom:0;'>📋 생보 위촉일정 조회 시스템</h2>", unsafe_allow_html=True)
        if st.session_state.user_dept == "ALL":
            st.markdown(f"<p style='color:#888; margin-top:4px;'>👤 {st.session_state.user_name}</p>", unsafe_allow_html=True)
        else:
            st.markdown(f"<p style='color:#888; margin-top:4px;'>👤 {st.session_state.user_name} | 🏢 {st.session_state.user_dept}</p>", unsafe_allow_html=True)
    with col3:
        if st.button("🔄 새로고침"):
            st.cache_data.clear()
            st.rerun()
        if st.button("로그아웃"):
            st.session_state.logged_in = False
            st.rerun()

    st.divider()

    # 데이터 로드 (환경에 따라 자동 분기)
    with st.spinner("데이터를 불러오는 중..."):
        if IS_LOCAL:
            sheets, error = load_excel_data(EXCEL_PATH)
        else:
            sheets, error = load_gsheets_data()

    if error:
        st.error(f"❌ {error}")
        if IS_LOCAL:
            st.info("ℹ️ 로컬 환경: `EXCEL_PATH` 변수를 실제 엑셀 파일 경로로 설정하세요.")
        else:
            st.info("ℹ️ 클라우드 환경: Streamlit Secrets 설정을 확인하세요.")
        st.stop()

    # 파일 정보 표시 생략 (사용자에게 불필요한 메타데이터 숨김)

    # 탭 생성
    if st.session_state.user_role == "admin":
        tabs = st.tabs(["🔍 위촉일정 검색", "📎 당월 위촉일정 안내", "⚠️ 보험사별 유의사항", "📝 대상자 추가요청", "👑 관리자 전용"])
        tab1, tab2, tab3, tab4, tab5 = tabs
    else:
        tabs = st.tabs(["🔍 위촉일정 검색", "📎 당월 위촉일정 안내", "⚠️ 보험사별 유의사항", "📝 대상자 추가요청"])
        tab1, tab2, tab3, tab4 = tabs

    # -------------------------------------------------------------
    # 탭 1: 검색
    # -------------------------------------------------------------
    with tab1:
        st.markdown("### 🔍 이름 또는 사원번호로 검색")
        
        # 검색 대상 선택 UI
        search_mode = st.radio("검색 범위", ["전체 보험사 검색", "특정 보험사 선택 검색"], horizontal=True, label_visibility="collapsed")
        
        selected_companies = None
        if search_mode == "특정 보험사 선택 검색":
            all_companies = [s for s in sheets.keys() if s != "notice"]
            selected_companies = st.multiselect(
                "검색할 보험사를 선택하세요", 
                options=all_companies, 
                placeholder="보험사 선택 (여러 개 선택 가능)"
            )
            
        st.markdown("<div style='margin-top: 10px;'></div>", unsafe_allow_html=True)
        col_search, col_btn = st.columns([5, 1])
        with col_search:
            keyword = st.text_input(
                label="검색어",
                placeholder="이름 또는 사원번호를 입력하세요...",
                label_visibility="collapsed"
            )
        with col_btn:
            search_clicked = st.button("검색", use_container_width=True, type="primary")

        if search_clicked or keyword:
            if search_mode == "특정 보험사 선택 검색" and not selected_companies:
                st.warning("검색할 보험사를 하나 이상 선택해 주세요.")
            elif not keyword.strip():
                st.warning("검색어를 입력해 주세요.")
            else:
                with st.spinner(f"'{keyword}' 검색 중..."):
                    results = search_all_sheets(sheets, keyword, st.session_state.user_dept, selected_companies)

                st.markdown(f"#### 📌 **{keyword}** 님의 위촉일정")

                if not results:
                    st.info("접수이력이 없습니다.")
                else:
                    result_df = pd.DataFrame(results)

                    for col in ["URL 동의기한", "위촉예정일", "미위촉시 다음접수일정"]:
                        if col in result_df.columns:
                            result_df[col] = result_df[col].apply(
                                lambda x: x[:10] if isinstance(x, str) and len(x) >= 10 else x
                            )

                    st.success(f"✅ {len(results)}개의 접수이력이 검색되었습니다.")
                    st.markdown("""
                    <div style="color: #d32f2f; font-size: 13.5px; background-color: #fff5f5; border-left: 4px solid #ef5350; padding: 12px 16px; margin-bottom: 16px; border-radius: 4px; line-height: 1.6;">
                        <b>1.</b> 코드 확인및 동의결과는 INS에서 해주시기 바랍니다.<br>
                        <b>2.</b> 재접수횟수를 참고하셔서 2회이상인 보험사가 URL 미제출일 경우 [대상자 추가요청]에 접수해주시기 바랍니다. (자동 재접수는 3회이상 하지 않습니다)<br>
                        <b>3.</b> 위촉예정일 익일 코드확인 및 사용 가능한 보험사 참조바랍니다.<br>
                        <span style="color: #c62828; font-size: 12.5px;">&nbsp;&nbsp;&nbsp;&nbsp;(신한라이프, DB생명, 흥국생명, 농협생명, KB라이프, IBK연금보험, 카디프생명)</span>
                    </div>
                    """, unsafe_allow_html=True)
                    st.dataframe(
                        result_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "보험사": st.column_config.TextColumn("🏦 보험사", width="medium"),
                            "▶": st.column_config.TextColumn("▶", width="small"),
                            "차수": st.column_config.TextColumn("차수", width="small"),
                            "URL 동의기한": st.column_config.TextColumn("📅 URL 동의기한", width="medium"),
                            "위촉예정일": st.column_config.TextColumn("📅 위촉예정일", width="medium"),
                            "미위촉시 다음접수일정": st.column_config.TextColumn("📅 미위촉시 다음접수일정", width="medium"),
                            "재접수횟수": st.column_config.TextColumn("🔄 재접수횟수", width="small")
                        }
                    )
        else:
            st.markdown("""
            <div style="text-align:center; padding:60px 0; color:#aaa;">
                <div style="font-size:48px;">🔎</div>
                <p style="font-size:16px; margin-top:12px;">검색어를 입력하고 검색 버튼을 누르세요.</p>
                <p style="font-size:13px;">이름 또는 사원번호로 전체 시트에서 최신 접수이력을 찾아드립니다.</p>
            </div>
            """, unsafe_allow_html=True)

    # -------------------------------------------------------------
    # 탭 2: 당월 위촉일정 안내
    # -------------------------------------------------------------
    with tab2:
        import glob
        
        pdf_path = os.path.join("assets", "schedule.pdf")
        has_pdf = os.path.exists(pdf_path)
        
        # 이미지 파일들 찾기 (schedule 로 시작하는 png, jpg, jpeg)
        image_files = []
        for ext in ('*.png', '*.jpg', '*.jpeg'):
            image_files.extend(glob.glob(os.path.join("assets", f"schedule{ext}")))
        image_files.sort()  # 이름순 정렬 (schedule1.jpg, schedule2.jpg ...)
        
        if has_pdf or image_files:
            col_t, col_b = st.columns([4, 1])
            with col_t:
                st.markdown("### 📎 당월 위촉일정 안내")
                if has_pdf and not image_files:
                    st.info("💡 클라우드 환경에서는 브라우저 보안 정책으로 인해 PDF 미리보기가 보이지 않을 수 있습니다. 우측 상단의 **[📥 PDF 다운로드]** 버튼을 눌러 확인해 주세요.")
            
            with col_b:
                if has_pdf:
                    st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
                    with open(pdf_path, "rb") as f:
                        pdf_bytes = f.read()
                    st.download_button(
                        label="📥 PDF 다운로드",
                        data=pdf_bytes,
                        file_name="위촉일정안내.pdf",
                        mime="application/pdf",
                        type="primary",
                        key="download_schedule",
                        use_container_width=True
                    )
            
            # 이미지 파일이 있으면 이미지들을 세로로 쭉 렌더링
            if image_files:
                st.markdown("<br>", unsafe_allow_html=True)
                for img_path in image_files:
                    st.image(img_path, use_container_width=True)
                    st.markdown("<br>", unsafe_allow_html=True)
            
            # 이미지가 없고 PDF만 있으면 기존의 iframe 미리보기 (로컬 환경 등 보이는 환경을 위해)
            elif has_pdf:
                import base64
                base64_pdf = base64.b64encode(pdf_bytes).decode('utf-8')
                pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="600" type="application/pdf" style="border:1px solid #ccc; border-radius:8px; margin-top:10px;"></iframe>'
                st.markdown(pdf_display, unsafe_allow_html=True)
        else:
            st.info("💡 등록된 당월 위촉일정 안내문이 없습니다. (assets 폴더에 schedule.pdf 또는 schedule1.jpg 파일을 추가해주세요)")

    # -------------------------------------------------------------
    # 탭 3: 보험사별 유의사항 (엑셀 notice 시트 연동)
    # -------------------------------------------------------------
    with tab3:
        st.markdown("### ⚠️ 보험사별 유의사항")
        
        notice_df = sheets.get("notice")
        
        if notice_df is not None and not notice_df.empty:
            if len(notice_df.columns) > 4:
                # A열(0), C열(2), E열(4) 추출
                n_df = notice_df.iloc[:, [0, 2, 4]].dropna(how='all')
                n_df.columns = ["대상", "키워드", "내용"]
                n_df = n_df.fillna("")
                
                # '대상'이 비어있는 행과 첫번째 행(보통 헤더일 가능성) 처리
                n_df = n_df[n_df["대상"].astype(str).str.strip() != ""]
                # 엑셀 헤더 행("보험사공통"이 아닌 그냥 "대상" 등)이 들어갔을 경우 제외
                n_df = n_df[~n_df["대상"].astype(str).str.contains("A열|대상|보험사명")]
                
                # 공통 유의사항 필터링
                common_df = n_df[n_df["대상"].astype(str).str.contains("공통")]
                if not common_df.empty:
                    st.markdown("#### 📌 공통 유의사항")
                    common_texts = []
                    for _, row in common_df.iterrows():
                        common_texts.append(f"**[{row['키워드']}]** {row['내용']}")
                    st.info("\n\n".join(common_texts))
                
                st.markdown("---")
                
                # 개별 보험사 유의사항 필터링
                company_df = n_df[~n_df["대상"].astype(str).str.contains("공통")]
                if not company_df.empty:
                    st.markdown("#### 🏢 보험사별 유의사항 검색")
                    company_list = company_df["대상"].unique().tolist()
                    selected_notice_company = st.selectbox(
                        "조회할 보험사를 선택하세요", 
                        options=company_list,
                        index=None,
                        placeholder="보험사를 선택해주세요"
                    )
                    
                    if selected_notice_company:
                        filtered_notice = company_df[company_df["대상"] == selected_notice_company]
                        company_texts = []
                        for _, row in filtered_notice.iterrows():
                            company_texts.append(f"**[{row['키워드']}]** {row['내용']}")
                        st.success("\n\n".join(company_texts))
            else:
                st.warning("엑셀 notice 시트의 형식이 맞지 않습니다. (최소 E열까지 데이터가 있어야 합니다.)")
        else:
            st.info("💡 엑셀 파일에 'notice' 시트가 없거나 비어 있습니다.")

    # -------------------------------------------------------------
    # 탭 4: 대상자 추가요청 (모든 사용자 접근 가능)
    # -------------------------------------------------------------
    with tab4:
        st.markdown("### 📝 대상자 추가 요청")
        st.info("""
💡 누락된 대상자가 있거나 새로 접수할 인원이 있다면 아래 폼을 작성해 제출해 주세요.
* **동양, 라이나, 처브**는 서면위촉 보험사이므로, 위촉서류 원본을 지원센터로 발송해주세요.
* **KB, 신한, 메트라이프, 하나생명**은 제출전 FA님께 고정링크 안내하시어 선제출 하시도록 안내해주세요.

🔗 **고정링크 바로가기**
* **메트라이프** → [📱 모바일버전](https://brand.metlife.co.kr/ga/appt/selectMoblieCert.do) / [💻 PC버전](https://metplus.metlife.co.kr)
* **KB라이프** → [👉 접속하기](http://id.kblife.co.kr/gafp/75400)
* **신한라이프** → [📱 모바일버전](https://ga.shinhanlife.co.kr:11043/mw/srhdjgp020m.mdv) / [💻 PC버전 (GA.com)](https://ga.shinhanlife.co.kr)
* **하나생명** → [📱 모바일접속](https://sales.hanalife.co.kr/ga/mo/MCM00007005M.do)
""")
        
        with st.form("request_form", clear_on_submit=True):
            col1, col2, col3 = st.columns(3)
            with col1:
                req_dept = st.text_input("🏢 소속명")
            with col2:
                req_emp_id = st.text_input("🆔 대상자 사번")
            with col3:
                req_name = st.text_input("👤 대상자 이름")
                
            # 최대 5개 선택 가능한 보험사 다중 선택 (서면위촉 보험사 3개 제외)
            insurance_options = [
                "ABL생명", "DB생명", "IBK연금", "iM라이프", "KB라이프생명", "KDB생명", 
                "NH농협생명", "교보생명", "메트라이프", "미래에셋", 
                "삼성생명", "신한라이프", "카디프생명", "푸본현대생명", 
                "하나생명", "한화생명", "흥국생명", "기타"
            ]
            req_companies = st.multiselect(
                "🏦 보험사 선택 (최대 5개)",
                options=insurance_options,
                max_selections=5,
                placeholder="보험사를 선택해주세요"
            )
                
            req_details = st.text_area("✍️ 요청 내용 (선택)")
            
            # 제출 버튼
            submitted_req = st.form_submit_button("제출하기", type="primary", use_container_width=True)
            
            if submitted_req:
                if not req_dept.strip() or not req_name.strip() or not req_companies:
                    st.error("⚠️ 소속명, 대상자 이름, 그리고 보험사(최소 1개)는 필수 입력 사항입니다.")
                else:
                    kst = timezone(timedelta(hours=9))
                    req_date = datetime.now(kst).strftime("%Y-%m-%d %H:%M:%S")
                    new_rows = []
                    
                    # 선택한 보험사별로 각각 행(Row) 생성
                    for company in req_companies:
                        new_rows.append({
                            "소속명": req_dept,
                            "대상자사번": req_emp_id,
                            "이름": req_name,
                            "보험사": company,
                            "요청내용": req_details,
                            "요청일시": req_date,
                            "상태": "신규"
                        })
                    
                    try:
                        import gspread
                        from google.oauth2.service_account import Credentials
                        creds_dict = dict(st.secrets["gcp_service_account"])
                        if "private_key" in creds_dict: creds_dict["private_key"] = creds_dict["private_key"].replace('\\n', '\n')
                        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
                        credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
                        client = gspread.authorize(credentials)
                        spreadsheet_url = "https://docs.google.com/spreadsheets/d/1jWwDgw0rEGeb0R4_d1UsfOnXsfwbOjtu9BBiPTZ16p0/edit?gid=0#gid=0"
                        sh = client.open_by_url(spreadsheet_url)
                        
                        try:
                            ws_req = sh.worksheet("requests")
                        except gspread.exceptions.WorksheetNotFound:
                            ws_req = sh.add_worksheet(title="requests", rows="100", cols="10")
                            ws_req.append_row(["소속명", "대상자사번", "이름", "보험사", "요청내용", "요청일시", "상태"])
                        
                        rows_to_append = [list(row.values()) for row in new_rows]
                        ws_req.append_rows(rows_to_append)
                        st.success(f"✅ 총 {len(new_rows)}건의 추가접수 요청이 구글 시트에 안전하게 전송되었습니다!")
                        st.cache_data.clear() # 관리자 탭에 즉시 반영되도록 캐시 초기화
                    except Exception as e:
                        st.error(f"요청 저장 중 오류가 발생했습니다: {e}")

    # -------------------------------------------------------------
    # 탭 5: 관리자 전용 (admin 권한만 보임)
    # -------------------------------------------------------------
    if st.session_state.user_role == "admin":
        with tab5:
            st.markdown("### 👑 관리자 전용: 접수된 추가 요청 목록")
            req_df_raw = sheets.get("requests")
            
            if req_df_raw is not None and not req_df_raw.empty:
                req_df = req_df_raw.copy()
                req_df.columns = req_df.iloc[0]
                req_df = req_df[1:].reset_index(drop=True)
                
                # '상태' 열이 없는 예외 상황 대비
                if "상태" not in req_df.columns:
                    req_df["상태"] = "신규"
                
                # 현재 '신규' (대기 중) 상태인 요청들 중에서만 보험사 목록 추출
                pending_companies = req_df[req_df["상태"] == "신규"]["보험사"].dropna().unique().tolist()
                
                selected_companies = st.multiselect(
                    "🔍 다운로드할 보험사 필터링 (현재 미처리 요청이 있는 회사만 표시됩니다)", 
                    options=pending_companies, 
                    default=[],
                    placeholder="보험사를 선택해주세요"
                )
                
                # 필터 적용
                if selected_companies:
                    filtered_df = req_df[req_df["보험사"].isin(selected_companies)]
                else:
                    filtered_df = req_df
                
                # 정렬 (최신순)
                if "요청일시" in filtered_df.columns:
                    filtered_df = filtered_df.sort_values(by="요청일시", ascending=False)
                
                # 상태별 데이터 분리
                pending_df = filtered_df[filtered_df["상태"] == "신규"]
                completed_df = filtered_df[filtered_df["상태"] == "다운로드 완료"]
                
                # 1. 미처리 신규 요청 영역
                st.markdown("#### 🚨 미처리 신규 요청")
                if pending_df.empty:
                    st.success("🎉 현재 대기 중인 신규 요청이 없습니다!")
                else:
                    st.dataframe(pending_df, use_container_width=True, hide_index=True)
                    
                    # 다운로드 버튼 클릭 시 상태 업데이트 콜백
                    def mark_as_downloaded(indices_to_update):
                        try:
                            import gspread
                            from google.oauth2.service_account import Credentials
                            creds_dict = dict(st.secrets["gcp_service_account"])
                            if "private_key" in creds_dict: creds_dict["private_key"] = creds_dict["private_key"].replace('\\n', '\n')
                            scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
                            credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
                            client = gspread.authorize(credentials)
                            sh = client.open_by_url("https://docs.google.com/spreadsheets/d/1jWwDgw0rEGeb0R4_d1UsfOnXsfwbOjtu9BBiPTZ16p0/edit?gid=0#gid=0")
                            ws_req = sh.worksheet("requests")
                            
                            cells_to_update = []
                            for idx in indices_to_update:
                                # pandas index는 0부터 시작, 헤더는 1행, 데이터 시작은 2행이므로 + 2
                                cells_to_update.append(gspread.Cell(row=idx + 2, col=7, value="다운로드 완료"))
                            ws_req.update_cells(cells_to_update)
                            st.cache_data.clear() # 캐시 갱신
                        except Exception as e:
                            print(f"다운로드 상태 변경 에러: {e}")
                            
                    # 다운로드할 데이터 준비 (신규 요청만)
                    csv_bytes = pending_df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
                    
                    # 다운로드 버튼 우측 하단 배치
                    col_blank, col_btn = st.columns([3, 1])
                    with col_btn:
                        st.download_button(
                            label="📥 신규 항목 마감 및 다운로드",
                            data=csv_bytes,
                            file_name=f"신규_추가요청목록_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv",
                            type="primary",
                            on_click=mark_as_downloaded,
                            args=(pending_df.index,),
                            use_container_width=True
                        )
                        
                st.markdown("---")
                
                # 2. 과거 처리 완료 내역 영역 (토글/Expander로 숨김 처리)
                with st.expander("📂 과거 처리(다운로드) 완료 내역 보기"):
                    if completed_df.empty:
                        st.info("아직 처리 완료된 내역이 없습니다.")
                    else:
                        def style_downloaded(row):
                            return ["color: #b0b0b0;"] * len(row)
                        
                        st.dataframe(completed_df.style.apply(style_downloaded, axis=1), use_container_width=True, hide_index=True)
            else:
                st.info("아직 접수된 대상자 추가 요청이 없습니다.")
                
            st.markdown("---")
            st.markdown("#### ☁️ 클라우드 서버 실시간 동기화")
            st.success("🤖 **자동 감시가 실행 중입니다.** 엑셀 파일을 저장(Ctrl+S)하면 약 10~20초 뒤 구글 시트로 자동 전송됩니다.")
            st.info("💡 (수동 동기화) 원하신다면 아래 버튼을 눌러 지금 즉시 강제로 전송할 수도 있습니다.")
            
            if st.button("🚀 엑셀 데이터 구글 시트로 덮어쓰기 (동기화 실행)", type="primary", use_container_width=True):
                with st.spinner("구글 스프레드시트로 데이터를 전송하고 있습니다... (데이터 양에 따라 최대 1분 소요)"):
                    try:
                        import sync_to_gsheets
                        
                        json_file = "dashboard-sync-503805-3ac60a044531.json"
                        gsheet_url = "https://docs.google.com/spreadsheets/d/1jWwDgw0rEGeb0R4_d1UsfOnXsfwbOjtu9BBiPTZ16p0/edit?gid=0#gid=0"
                        
                        success, msg = sync_to_gsheets.sync_excel_to_gsheets(EXCEL_PATH, json_file, gsheet_url)
                        if success:
                            st.success(msg)
                            st.balloons()
                        else:
                            st.error(msg)
                    except Exception as e:
                        st.error(f"동기화 실행 중 오류가 발생했습니다: {str(e)}")


# ─────────────────────────────────────────────────────────────
#  페이지 설정 및 진입점
# ─────────────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="생보 위촉일정 조회 시스템",
        page_icon="📋",
        layout="wide",
        initial_sidebar_state="collapsed"
    )

    # 커스텀 CSS
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700&display=swap');
        html, body, [class*="css"] {
            font-family: 'Noto Sans KR', sans-serif;
        }
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        .stButton > button {
            border-radius: 8px;
            font-weight: 500;
        }
        .stButton > button[kind="primary"] {
            background-color: #4A90D9;
            border-color: #4A90D9;
        }
        .stDataFrame {
            border-radius: 8px;
        }
        div[data-testid="stForm"] {
            border: 1px solid #e0e0e0;
            border-radius: 12px;
            padding: 24px;
            background: white;
            box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        }
        /* 탭 버튼 직관적으로 스타일링 */
        button[data-baseweb="tab"] {
            background-color: #f8f9fa;
            border: 1px solid #e9ecef;
            border-bottom: none;
            border-radius: 8px 8px 0 0;
            padding: 10px 24px !important;
            margin-right: 4px;
            box-shadow: inset 0 -2px 0 0 #e9ecef;
        }
        button[data-baseweb="tab"]:hover {
            background-color: #e9ecef;
        }
        button[data-baseweb="tab"][aria-selected="true"] {
            background-color: white;
            border: 1px solid #dee2e6;
            border-bottom: none;
            box-shadow: inset 0 3px 0 0 #4A90D9;
            font-weight: bold;
        }
    </style>
    """, unsafe_allow_html=True)

    init_session()

    if not st.session_state.logged_in:
        render_login()
    else:
        render_dashboard()


if __name__ == "__main__":
    main()
